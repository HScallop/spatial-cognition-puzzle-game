import glob
import numpy as np
from decimal import Decimal
from sklearn.cluster import KMeans
import matplotlib.pyplot as plt
import openpyxl as pyxl


def type_encode(type_str):
    rstr=''
    for i in range(10):
        rstr=rstr+str(type_str.count(str(i)))
    return rstr

def question_type(lines):
    types=[]
    types_code=[]
    for line in lines:
        types.append(line[10::4])
    for i in types:
        types_code.append(type_encode(i))
    return types_code

def get_correctness(lines):
    rlist=[]
    for line in lines:
        rlist.append(line[0])
    return rlist

def get_time(lines):
    rtime=[]
    for line in lines:
        rtime.append((float(line[1:5])))
    #print('rtime')
    #print(rtime)
    return rtime

def sumstr(s):
    temp=0
    for char in s:
        temp+=int(char)
    return temp

def non0(s):
    return len(s)-s.count('0')

def code(line_list):
    rlist=[]
    for i in range(len(line_list)):
        tempstr=''
        tempstr=str(sumstr(line_list[i]))+str(sumstr(line_list[i][0:5]))+str(sumstr(line_list[i][5:]))+str(non0(line_list[i][0:5]))+str(non0(line_list[i][5:]))
        rlist.append(tempstr)
    return rlist

#read all data files
files=glob.glob('data30/B/*.txt')
print(files)
temp=0
question_list=[]
question_dic={}
question_data={}
with open(files[0],'r') as f:
    lines=f.readlines()
    lines=[line.strip('\n') for line in lines]
    question_list=question_type(lines)
    question_dic={questions:[] for questions in question_list}
    question_data={questions:[] for questions in question_list}
for name in files:
    with open(name,'r')as f:
        lines=f.readlines()
        lines=[line.strip('\n') for line in lines]
        question_list=question_type(lines)
        question_correctness=get_correctness(lines)
        question_time=get_time(lines)
        for i in range(len(question_time)):
            try:
                question_dic[question_list[i]].append(question_correctness[i])
                question_dic[question_list[i]].append(question_time[i])
            except:
                print(name)
                print('e')
for i in question_dic:
    count=0
    for j in range(0,len(question_dic[i]),2):
        if question_dic[i][j]=='0':
            continue
        temp+=int(question_dic[i][j])-1
        count+=1
    if count!=0:
        question_data[i].append(temp/count)
    temp=0
    for j in range(1,len(question_dic[i]),2):
        if question_dic[i][j-1]=='0':
            continue
        temp+=int(question_dic[i][j])
    if count!=0:
        question_data[i].append(temp/count)
    temp=0

#print(len(question_dic['0010211210']))
#print(question_list)
#print(len(question_list))
cor_time_list=[]
print('qdata')
print(question_data)
for i in question_data:
    cor_time_list.append(question_data[i])
#print(cor_time_list)

#to encode question type to the required ones
question_list_code=code(list(question_data.keys()))
print(question_list_code)
print(len(question_list_code))

#print it to .xlxs file
wb=pyxl.Workbook()
ws=wb.active
for col in ws.iter_cols(min_row=1, min_col=1, max_col=1, max_row=len(question_list_code)):
    for cell, code in zip(col, question_list_code):
        cell.value=code

for col in ws.iter_cols(min_row=1, min_col=2, max_col=2, max_row=len(question_list_code)):
    for cell, acc_time in zip(col, question_data.values()):
        cell.value=acc_time[0]

for col in ws.iter_cols(min_row=1, min_col=3, max_col=3, max_row=len(question_list_code)):
    for cell, acc_time in zip(col, question_data.values()):
        cell.value=acc_time[1]

wb.save('codeB.xlsx')

'''
for i in range(len(question_list)):
    print(question_list[i])
    print(cor_time_list[i])
'''
#print(len(cor_time_list))
for i in range(len(cor_time_list)):
    if cor_time_list[i][1]==0.0:
        print(i)
X=np.array(cor_time_list, dtype=object)

km=KMeans(n_clusters=10,init='random',n_init=10,max_iter=300,tol=1e-04,random_state=0)
y_km=km.fit_predict(X)
#print("y_km")
#print(y_km)
#print("center")
#print(km.cluster_centers_)
ccluster=[]
for i in range(6):
    ccluster.append(km.cluster_centers_[i][0]/km.cluster_centers_[i][1])
#print("ccluster")
#print(ccluster)
for i in range(6):
    for j in range(len(y_km)):
        if(i==y_km[j]):
            #print(i)
            #print(question_list[j])
            print(' ')
plt.scatter(
    X[y_km == 0, 0], X[y_km == 0, 1],
    s=50, c='lightgreen',
    marker='s', edgecolor='black',
    label='cluster 1'
)

plt.scatter(
    X[y_km == 1, 0], X[y_km == 1, 1],
    s=50, c='orange',
    marker='o', edgecolor='black',
    label='cluster 2'
)

plt.scatter(
    X[y_km == 2, 0], X[y_km == 2, 1],
    s=50, c='lightblue',
    marker='v', edgecolor='black',
    label='cluster 3'
)

plt.scatter(
    X[y_km == 3, 0], X[y_km == 3, 1],
    s=50, c='red',
    marker='d', edgecolor='black',
    label='cluster 4'
)
plt.scatter(
    X[y_km==4,0],X[y_km==4,1],
    s=50,c='yellow',
    marker='.',edgecolor='black',
    label='cluster 5'
)
plt.scatter(
    X[y_km==5,0],X[y_km==5,1],
    s=50,c='purple',
    marker='d',edgecolor='black',
    label='cluster 6'
)
plt.scatter(
    X[y_km==6,0],X[y_km==6,1],
    s=50,c='blue',
    marker='>',edgecolor='black',
    label='cluster 7'
)
plt.scatter(
    X[y_km==7,0],X[y_km==7,1],
    s=50,c='pink',
    marker='<',edgecolor='black',
    label='cluster 8'
)
plt.scatter(
    X[y_km==8,0],X[y_km==8,1],
    s=50,c='black',
    marker='p',edgecolor='black',
    label='cluster 9'
)
plt.scatter(
    X[y_km==9,0],X[y_km==9,1],
    s=50,c='gray',
    marker='8',edgecolor='black',
    label='cluster 10'
)
plt.scatter(
    km.cluster_centers_[:, 0], km.cluster_centers_[:, 1],
    s=250, marker='*',
    c='red', edgecolor='black',
    label='centroids'
)

plt.legend(scatterpoints=1)
plt.grid()
plt.show()

'''
y_lm_list=[5,4,0,3,2,1]
with open('result.txt','a') as f:
    for i in range(len(question_list)):
        print(dir(y_km[i]),file='f')
        print('\n')
        print(dir(question_list[i]),file='f')
'''






